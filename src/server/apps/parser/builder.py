import os
import datetime

from django.core.mail import EmailMessage
from django.conf import settings

import openpyxl
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink


class ExcelEmailSender:
    def __init__(self, data):
        self.data = data

    def create_excel_file(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        # Записываем заголовки
        headers = [
            '№', 'Ссылка', 'Статус', 'Субъект', 'Заказчик',
            'Размещено', 'Обновлено', 'Окончание подачи заявки', 'Начальная цена'
        ]
        ws.append(headers)

        # Записываем данные
        for item in self.data:
            ws.append([
                item['number'], item['url'], item['status'], item['subject'], item['customer'],
                item['start_date'], item['change_date'], item['stop_date'], item['price']
            ])

        # Делаем границы и выравнивание
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.column == 2:  # преобразуем ссылку в кликабельную
                    cell_hyperlink = Hyperlink(ref=cell.value, target=cell.value)
                    cell.value = None
                    cell.hyperlink = cell_hyperlink

                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = Border(left=Side(border_style="thin"),
                                     right=Side(border_style="thin"),
                                     top=Side(border_style="thin"),
                                     bottom=Side(border_style="thin"))

        # Регулируем ширину колонок для лучшего отображения
        for col in ws.columns:
            max_len = max(len(str(cell.value)) for cell in col)
            column = get_column_letter(col[0].column)
            adjusted_width = (max_len + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        filename = f'аукционы_{datetime.date.today().strftime("%d-%m-%Y")}.xlsx'
        wb.save(filename)
        return filename

    def send_email(self, recipient_email):
        subject = f'Аукционы-за-{datetime.date.today().strftime("%d-%m-%Y")}'
        body = 'Все аукционы находятся в файле.'

        filename = self.create_excel_file()

        email = EmailMessage(
            subject,
            body,
            settings.EMAIL_HOST_USER,
            [recipient_email],
        )
        email.attach_file(filename)
        email.send()
        os.remove(filename)
